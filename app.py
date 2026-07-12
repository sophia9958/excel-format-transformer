import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import difflib
import json
import base64
import openpyxl
import streamlit.components.v1 as components

# --- 1. 辅助工具函数 ---
def is_duration_col(col_name, custom_time_keywords=None):
    """智能识别时间/时长列"""
    include_k = custom_time_keywords if custom_time_keywords else ['总', '时长', '时间', '片长', '总长', '总片长', 'Duration', 'time']
    exclude_k = ['上线', '日期', '发布', '开播', '年度', '更新', 'date', 'day', '总集数', '总片数', '总集', 'ID', '号', 'No']
    is_inc = any(k in str(col_name) for k in include_k)
    is_exc = any(k in str(col_name) for k in exclude_k)
    return is_inc and not is_exc

def parse_time_logic(val):
    """支持超过 24 小时的时间数据强制解析"""
    if pd.isna(val) or val == "": return ""
    v_str = str(val).strip()
    if ':' in v_str:
        parts = v_str.split(':')
        try:
            if len(parts) == 3:
                return (int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])) / 86400.0
            elif len(parts) == 2:
                return (int(parts[0]) * 60 + int(parts[1])) / 84600.0
        except: pass
    try: return float(v_str)
    except: return v_str

def find_headers(df, custom_keywords=None):
    """
    智能侦测表头所在行：扩大到 200 行，适应极其变态的超长表头模版
    """
    best_idx, max_m = 0, -1
    core = custom_keywords if custom_keywords else {'编码', '名称', '日期', '类型', '状态', '标题', '时间', 'ID', 'Name', 'Date', 'Type'}
    
    for idx in range(min(200, len(df))):
        row = [str(x).strip() for x in df.iloc[idx].values if pd.notna(x)]
        m = sum(3 if any(c in v for c in core) else (0.8 if len(v)>1 else 0) for v in row)
        
        if len(row) > 0:
            num_count = sum(1 for v in row if v.replace('.','',1).isdigit())
            if num_count / len(row) > 0.5: m -= 5
                
        if m > max_m: max_m, best_idx = m, idx
        
    h = [str(x).strip() if pd.notna(x) and str(x).strip() != '' else f"Unnamed_{i}" for i, x in enumerate(df.iloc[best_idx].values)]
    return best_idx, h

# --- 2. 页面基本配置 ---
st.set_page_config(page_title="万能 Excel 表头提取与排版助手", layout="wide", page_icon="🔀")

# --- 3. 顶部说明区 ---
st.title("🔀 万能 Excel 表头提取与排版助手")

st.markdown("""
✅ **无损保留模版**：完美保留表 B 中的所有前置说明文字、合并单元格及底色格式。
✅ **智能 VLOOKUP**：若表 B 中已预填“编码/名称”，系统会自动根据关键字段去表 A 精准捞取数据，防止多 Sheet 数据乱跑串台！
✅ **自动时间进位**：考虑到影视/工时时长会超过 24 小时，系统强转 `[h]:mm:ss`。
---
* **自定义填充**：A表没有的字段，可通过左侧第 1 点默认填满。
* **无表头预警**：红灯预警空表头，拒绝瞎匹配。
* **求助与排错**：若点击运行没反应，可点击网页右上角 `⋮` 选择 **Clear cache**，若报错请复制底部日志反馈给 **nolinda@126.com**。
""")

# --- 4. 侧边栏：1. 个性化默认值填充 ---
st.sidebar.header("⚙️ 1. 个性化默认值填充")
st.sidebar.info("👉 **用法：** 表B字段名=你要填的内容。未设置的字段为空。")
today_str = datetime.datetime.now().strftime("%y%m%d")
custom_defaults_text = st.sidebar.text_area("输入填充规则：", placeholder=f"更新日期={today_str}\n是否成品=是", height=150)

custom_defaults = {}
if custom_defaults_text.strip():
    for line in custom_defaults_text.split('\n'):
        if '=' in line:
            parts = line.split('=')
            if len(parts) == 2: custom_defaults[parts[0].strip()] = parts[1].strip()

# --- 5. 侧边栏：2. 手动对号修正 ---
st.sidebar.markdown("---")
st.sidebar.header("🛠️ 2. 手动对号修正")
st.sidebar.info("👉 **用法：** 表B字段名=表A列号。例如：UMAI=58")
manual_map_text = st.sidebar.text_area("输入映射规则（不区分大小写）：", placeholder="许可证=10", height=120)

manual_map_config = {}
if manual_map_text.strip():
    for line in manual_map_text.split('\n'):
        if '=' in line:
            parts = line.split('=')
            if len(parts) == 2:
                try: manual_map_config[parts[0].strip().lower()] = int(parts[1].strip()) - 1
                except ValueError: pass

# --- 6. 主界面：文件上传 ---
col_u1, col_u2 = st.columns(2)
with col_u1:
    raw_file = st.file_uploader("📂 上传【表A：数据源】(系统导出表)", type=["csv", "xlsx", "xls"])
with col_u2:
    template_file = st.file_uploader("📋 上传【表B：目标模板】(格式模版)", type=["xlsx", "xls"])

# ==================== 7. 读取表A与无头预警 ====================
df_raw = None
if raw_file:
    try:
        ext = raw_file.name.split('.')[-1].lower()
        if ext == 'csv': df_raw = pd.read_csv(raw_file, dtype=str).fillna("")
        else: df_raw = pd.read_excel(raw_file, dtype=str).fillna("")
        
        st.success(f"✅ 表A 读取成功！共包含 {len(df_raw.columns)} 列，{len(df_raw)} 条数据。")
        
        unnamed = []
        for i, col in enumerate(df_raw.columns, 1):
            if "Unnamed" in str(col) or str(col).strip() == "":
                sample = [str(x).strip() for x in df_raw.iloc[:, i-1].tolist() if str(x).strip() != ""]
                unnamed.append((i, "、".join(sample[:3]) if sample else "全空"))
        if unnamed:
            st.warning("🚨 **表A 发现无名列！** 请根据预览在左侧手动对号：")
            for idx, pre in unnamed: st.write(f"👉 第 `{idx}` 列 ➔ 预览: `{pre}`")
    except Exception as e: 
        st.error(f"读取表A失败: {e}")
        st.stop()

# ==================== 8. 侧边栏：3. 自查与特征词 ====================
selected_time_cols = []
force_header_config = {}

st.sidebar.markdown("---")
st.sidebar.header("🔍 3. 进阶特征词控制")

header_keywords_input = st.sidebar.text_input(
    "🧠 表头雷达特征词 (逗号隔开)",
    value="编码,名称,日期,类型,状态,标题,时间,片单,导演,演员,ID,Name,Date,Type"
)
custom_header_keywords = set(x.strip() for x in header_keywords_input.replace("，", ",").split(",") if x.strip())

time_keywords_input = st.sidebar.text_input(
    "⏳ 时长转换特征词 (逗号隔开)",
    value="总,时长,时间,片长,总长,总片长,Duration,time"
)
custom_time_keywords = [x.strip() for x in time_keywords_input.replace("，", ",").split(",") if x.strip()]

if template_file:
    # 只需读取表B页签名称和预览
    xls_tpl = pd.ExcelFile(template_file)
    for s_name in xls_tpl.sheet_names:
        force_header_config[s_name] = st.sidebar.number_input(f"『{s_name}』表头在第几行(0为自动)", min_value=0, max_value=200, value=0, key=f"r_{s_name}")
    
    st.sidebar.markdown("---")
    st.sidebar.subheader("⏳ 时间格式字段确认")
    
    all_time_cols_in_template = []
    for s_name in xls_tpl.sheet_names:
        try:
            df_preview = pd.read_excel(xls_tpl, sheet_name=s_name, header=None, nrows=200)
            if not df_preview.empty:
                _, hs = find_headers(df_preview, custom_header_keywords)
                all_time_cols_in_template.extend([h for h in hs if is_duration_col(h, custom_time_keywords)])
        except: pass
    all_time_cols_in_template = list(set(all_time_cols_in_template))
    
    selected_time_cols = st.sidebar.multiselect("以下字段执行 [h]:mm:ss 累加", all_time_cols_in_template, default=all_time_cols_in_template)

# ==================== 9. 原生修改模式与 VLOOKUP ====================
if df_raw is not None and template_file:
    output = BytesIO()
    final_reports = {}
    debug_log = {"诊断": {}}

    # 使用 openpyxl 加载原始工作簿（100% 保护格式和排版）
    wb = openpyxl.load_workbook(template_file)

    for s_name in wb.sheetnames:
        ws = wb[s_name]
        
        # 将 sheet 前 200 行转为 dataframe 用于定位表头
        data_preview = []
        max_r = min(200, ws.max_row)
        for r in range(1, max_r + 1):
            data_preview.append([cell.value for cell in ws[r]])
        df_tpl_meta = pd.DataFrame(data_preview)
        
        if df_tpl_meta.empty: continue
        
        f_h = force_header_config.get(s_name, 0)
        if f_h > 0:
            h_idx = f_h - 1
            headers = [str(x).strip() if pd.notna(x) and str(x).strip()!='' else f"Unnamed_{i}" for i, x in enumerate(df_tpl_meta.iloc[h_idx].values)]
        else:
            h_idx, headers = find_headers(df_tpl_meta, custom_header_keywords)
            
        raw_cols = df_raw.columns.tolist()
        mapped_cols = {} # {b_col_idx: a_col_idx}
        report = []
        
        # 1. 执行表头相似度匹配
        for b_idx, col_name in enumerate(headers):
            if col_name.startswith("Unnamed_") or not col_name.strip(): continue
            a_idx, status = None, "empty"
            
            if col_name.lower() in manual_map_config:
                a_idx = manual_map_config[col_name.lower()]; status = "ok"
            else:
                m = difflib.get_close_matches(col_name, raw_cols, n=1, cutoff=0.4)
                if m: a_idx = raw_cols.index(m[0]); status = "ok"
                
            if status == "ok" and a_idx < len(raw_cols):
                mapped_cols[b_idx] = a_idx
                report.append({"b": b_idx+1, "bn": col_name, "a": a_idx+1, "an": raw_cols[a_idx], "s": "ok"})
            else:
                fill = custom_defaults.get(col_name, "")
                report.append({"b": b_idx+1, "bn": col_name, "f": fill, "s": "fill" if fill else "empty"})
                
        # 2. 侦测：是执行 VLOOKUP 还是直接向下覆盖追加？
        has_data = False
        if ws.max_row > h_idx + 1:
            for cell in ws[h_idx + 2]:
                if cell.value is not None and str(cell.value).strip() != "":
                    has_data = True
                    break
                    
        if has_data and mapped_cols:
            # === VLOOKUP 模式（表B已有待匹配的影片行）===
            # 寻找用来 VLOOKUP 的基准主键（如 编码 或 名称）
            key_b_idx = list(mapped_cols.keys())[0]
            for b_c in mapped_cols.keys():
                if any(k in headers[b_c].lower() for k in ['编码', 'id', '名称', '标题', '片单']):
                    key_b_idx = b_c
                    break
            key_a_idx = mapped_cols[key_b_idx]
            
            # 构建 A 表哈希字典提升速度
            a_lookup = {}
            for i in range(len(df_raw)):
                k_val = str(df_raw.iloc[i, key_a_idx]).strip()
                if k_val: a_lookup[k_val] = df_raw.iloc[i]
                
            # 逐行对比填写
            for r in range(h_idx + 2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=key_b_idx + 1).value
                k_val = str(cell_val).strip() if cell_val is not None else ""
                
                if k_val in a_lookup:
                    a_row = a_lookup[k_val]
                    for b_c, a_c in mapped_cols.items():
                        if b_c == key_b_idx: continue # 不覆盖用来匹配的基准词
                        val = a_row.iloc[a_c]
                        if headers[b_c] in selected_time_cols: val = parse_time_logic(val)
                        
                        cell = ws.cell(row=r, column=b_c + 1)
                        cell.value = val
                        if isinstance(val, (int, float)): cell.number_format = '[h]:mm:ss'
                        
                    for b_c, col_name in enumerate(headers):
                        if col_name in custom_defaults and b_c not in mapped_cols:
                            ws.cell(row=r, column=b_c + 1).value = custom_defaults[col_name]
        else:
            # === 纯净追加模式（表B下方全空，原样照搬A表全部内容） ===
            for i in range(len(df_raw)):
                r = h_idx + 2 + i
                for b_c, a_c in mapped_cols.items():
                    val = df_raw.iloc[i, a_c]
                    if headers[b_c] in selected_time_cols: val = parse_time_logic(val)
                    
                    cell = ws.cell(row=r, column=b_c + 1)
                    cell.value = val
                    if isinstance(val, (int, float)): cell.number_format = '[h]:mm:ss'
                        
                for b_c, col_name in enumerate(headers):
                    if col_name in custom_defaults and b_c not in mapped_cols:
                        ws.cell(row=r, column=b_c + 1).value = custom_defaults[col_name]
                        
        final_reports[s_name] = report
        debug_log["诊断"][s_name] = {"识别行": h_idx+1, "模式": "VLOOKUP匹配" if has_data else "直接提取写入"}

    # 保存工作簿
    wb.save(output)

    st.markdown("---")
    st.subheader("📊 数据对齐与填充看板")
    for s_name, report in final_reports.items():
        ok_count = sum(1 for x in report if x['s'] == 'ok')
        mode_str = debug_log["诊断"][s_name]["模式"]
        with st.expander(f"📁 『{s_name}』 [{mode_str}] (匹配字段: {ok_count} | 填充: {len(report)-ok_count})"):
            c1, c2 = st.columns(2)
            with c1:
                st.write("🟢 **表A提取得出**")
                for x in [item for item in report if item['s'] == 'ok']: st.write(f"A第`{x['a']}`列({x['an']}) ➔ B第`{x['b']}`列({x['bn']})")
            with c2:
                st.write("🟡 **自定义填充或留空**")
                for x in [item for item in report if item['s'] != 'ok']:
                    val_text = f"填: `{x['f']}`" if x['f'] else "留空"
                    st.write(f"B第`{x['b']}`列({x['bn']}) ➔ {val_text}")

    st.success("🎉 处理完成！全部排版已保护，VLOOKUP已自动执行。")
    st.download_button("📥 下载完美格式转换表", data=output.getvalue(), file_name=f"交付表_{today_str}.xlsx")
    
    st.markdown("---")
    with st.expander("🛠️ 排错专区"):
        log_json = json.dumps(debug_log, ensure_ascii=False, indent=2)
        st.code(log_json, language="json")
else:
    st.info("💡 请在上方分别上传【表A：包含数据的原始表】、【表B：排好版的格式模版】。")
